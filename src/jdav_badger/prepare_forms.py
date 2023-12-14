import csv
import os
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side

# input csv file
INPUT_FILE = 'JLMarken.csv'

# template for output excel file
TEMPLATE_NAME = 'Jugendleitermarken Bestellliste'
TEMPLATE_FILE = TEMPLATE_NAME + '.xlsx'

# column indexes in input csv file
GIVEN_NAME_COLUMN = 0
FAMILY_NAME_COLUMN = 1
CHAPTER_COLUMN = 2
BADGE_COLUMN = 3

# row to insert data at
INSERT_ROW = 4

# folder to collect output files in
OUTPUT_FOLDER = 'out'

# map badge status to full word
badgeMap = {
    'J': 'Ja',
    'B': 'Bedingt',
    'N': 'Nein',
    '': 'Nein',
}

result = {}

print('Daten werden eingelesen...')

with open(INPUT_FILE, 'r', encoding='cp1252') as file:
    reader = csv.reader(file, delimiter=';')

    # skip header
    next(reader, None)

    for row in reader:
        stripped = [s.strip() for s in row]
        chapter = stripped[CHAPTER_COLUMN].replace('/', '_')
        badge = badgeMap[stripped[BADGE_COLUMN].upper()]

        # copy data
        data = [
            stripped[GIVEN_NAME_COLUMN],
            stripped[FAMILY_NAME_COLUMN],
            badge,
        ]

        # store data by chapter name
        if chapter in result:
            result[chapter].append(data)
        else:
            result[chapter] = [data]

# make sure output directory exists
if not os.path.isdir(OUTPUT_FOLDER):
    os.mkdir(OUTPUT_FOLDER)

# create border
sideThin = Side(style=None, color='FF000000', border_style='thin')
border = Border(left=sideThin, right=sideThin, bottom=sideThin, top=sideThin)

print('Dateien werden erstellt...')

for chapter in result.keys():
    print('Aktuelle Sektion: ' + chapter, end="\x1b[1K\r")

    # open worksheet
    workbook = load_workbook(TEMPLATE_FILE)
    worksheet = workbook.active

    # insert sufficient rows
    worksheet.insert_rows(INSERT_ROW, amount=len(result[chapter]))

    # write cells
    for row_index, row_data in enumerate(result[chapter], start=INSERT_ROW):
        for col_index, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index).value = cell_value

        for i in range(8):
            worksheet.cell(row=row_index, column=i+1).border = border

    # save result to output folder
    workbook.save(OUTPUT_FOLDER + '/' + TEMPLATE_NAME + ' ' + chapter + '.xlsx')

print('Vorgang abgeschlossen!')
